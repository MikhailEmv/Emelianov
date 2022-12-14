import csv
import datetime

import numpy as np
from matplotlib import pyplot as plt

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

from jinja2 import Environment, FileSystemLoader
import pdfkit

from unittest import TestCase


class VacancyTests(TestCase):
    dictionary = \
        {
            'name': 'vacancies.csv',
            'salary_from': 100,
            'salary_to': 200,
            'salary_currency': 150,
            'area_name': 'area',
            'published_at': '2022-06-14T11:44:58+0300'
        }

    def test_vacancy_type(self):
        self.assertEqual(type(Vacancy(self.dictionary)).__name__, 'Vacancy')

    def test_vacancy_area_name(self):
        self.assertEqual(Vacancy(self.dictionary).area_name, 'area_name')

    def test_vacancy_name(self):
        self.assertEqual(Vacancy(self.dictionary).area_name, 'vacancies.csv')

    def test_vacancy_salary(self):
        self.assertEqual(Vacancy(self.dictionary).area_name, 22500)


class DataSetTests(TestCase):
    def test_dataset_type(self):
        self.assertEqual(type(DataSet('vacancies_big.csv', ['information'])).__name__, 'DataSet')

    def test_dataset_type2(self):
        self.assertEqual(type(DataSet('vacancies_big.csv',
                                      ['information']).vacancies_objects).__name__, 'list')

    def test_dataset_name(self):
        self.assertEqual(DataSet('vacancies_big.csv', ['information']).file_name, 'vacancies_big.csv')


class Vacancy:
    """Класс вакансии

    Attributes:
        name (str): Имя
        salary (int): Зарплата
        area_name (str): Регион
        published_at (str): Дата публикации
    """
    currency_ratio = \
        {
            "AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055,
        }

    def __init__(self, object_vacancy) -> None:
        """Инициализирует класс вакансии из словаря вакансии

        param object_vacancy: Словарь вакансии

        >>> dictionary = {'name': 'vacancies.csv', 'salary_from': 100 ,'salary_to': 200,
        >>> 'salary_currency': 150, 'area_name': 'area', 'published_at': '2022-06-14T11:44:58+0300'}
        >>> type(Vacancy(dictionary)).__name__
        'Vacancy'
        >>> Vacancy(dictionary).area_name
        'area'
        >>> Vacancy(dictionary).name
        'vacancies.csv'
        >>> Vacancy(dictionary).salary
        22500
        """
        self.name = object_vacancy['name']
        salary_from = int((float(("".join(object_vacancy['salary_from'].split())))))
        salary_to = int((float(("".join(object_vacancy['salary_to'].split())))))
        self.salary = (salary_from + salary_to) * self.currency_ratio[object_vacancy['salary_currency']] // 2
        self.area_name = object_vacancy['area_name']
        # str(parser.parse(dict_vacancy['published_at']).date())
        # '.'.join(str(datetime.datetime.strptime(dict_vacancy['published_at'],
        # '%Y-%m-%dT%H:%M:%S%z').date()).split('-'))
        self.published_at = datetime.datetime.strptime(object_vacancy['published_at'], '%Y-%m-%dT%H:%M:%S%z')


class DataSet:
    """Класс множества данных (dataset)

    Attributes:
        vacancies_objects (str): Лист вакансий
        file_name (str): Название файла
    """
    def __init__(self, file_name: str, vacancies_objects: list) -> None:
        """Конструктор класса DataSet

        param file_name: Название файла
        :param vacancies_objects: Лист вакансий

        >>> type(DataSet('vacancies_big.csv', ['information'])).__name__
        'DataSet'
        >>> type(DataSet('vacancies_big.csv', ['information']).vacancies_objects).__name__
        'list'
        >>> DataSet('vacancies_big.csv', ['information']).file_name
        'vacancies_big.csv'
        """
        self.vacancies_objects = vacancies_objects
        self.file_name = file_name

    def __csv_reader(self) -> tuple:
        """Приватный метод класса DataSet, выполняющий функции чтения файла

        :return: tuple из двух листов
        """
        headlines, vacancies = list(), list()
        with open(self.file_name, encoding='utf-8-sig') as file:
            vacancies_list, counter = csv.reader(file, delimiter=','), 0
            for line in vacancies_list:
                if counter == 0:
                    counter += 1
                    headlines = line
                else:
                    if '' in line or len(line) != len(headlines):
                        continue
                    vacancies.append(line)
        if len(headlines) == 0:
            print('Пустой файл')
            exit()
        if len(vacancies) == 0:
            print('Нет данных')
            exit()
        return vacancies, headlines

    @staticmethod
    def __csv_filer(reader: tuple, headlines: list) -> list:
        """Приватный статический метод для фильтрации

        :param reader: принимаем reader типа tuple - уже прочтенный файл
        :param headlines: принимаем лист-шапку

        :return: лист отфильтрованных вакансий
        """
        vacancies_list = list()
        for line in reader:
            current_dictionary = dict()
            for i in range(len(line)):
                current_dictionary[headlines[i]] = line[i]
            vacancies_list.append(Vacancy(current_dictionary))
        return vacancies_list

    def put_vacancies(self) -> None:
        """Вкладываем вакансии в новый объект

        :return: nothing
        """
        (vacancies, headlines) = self.__csv_reader()
        self.vacancies_objects = self.__csv_filer(vacancies, headlines)


class CustomTuple:
    """Класс CustomTuple

    Attributes:
        totalSalary (int): Количество цельных зарплат
        count (int): Их общее количество
    """
    def __init__(self, full_salary=0, counter=0) -> None:
        """Конструктор класса CustomTuple

        :param full_salary: Количество цельных зарплат
        :param counter: Их общее количество

        >>> CustomTuple().totalSalary
        0
        >>> CustomTuple().count
        0
        """
        self.totalSalary = full_salary
        self.count = counter


class InputConnect:
    """Класс ввода с консоли и вывода таблицы в консоль

    Attributes:
        list_of_all_dictionaries (list): Лист всех словарей с конкретизированной статистикой
    """
    years_stats, cities_stats, vacancy_stats = dict(), dict(), dict()

    def __init__(self):
        """Конструктор класса InputConnect
        """
        self.list_of_all_dictionaries = list()

    def start_entering(self) -> None:
        """Метод для ввода необходимых данных от пользователя
        Attributes:
            file_name (list): Имя файла
            profession (str): Имя профессии
            word_for_choice (str): Слово для выборки и нужд пользователя

        :return: nothing
        """
        self.file_name = input('Введите название файла: ')
        self.profession = input('Введите наименование профессии: ')
        self.word_for_choice = input('Введите "Вакансии" или "Статистика": ')
        self.cities_count = 0
        self.report = Report()

    def count_vacancies(self, vacancies_list: list) -> None:
        """Метод подсчета вакансий и их распределение по словарям

        :param vacancies_list: Лист всех вакансий

        :return: nothing
        """
        for vacancy in vacancies_list:
            self.cities_count += 1
            current_year = int(vacancy.published_at.year)
            if current_year not in self.years_stats.keys():
                self.years_stats[current_year] = CustomTuple(vacancy.salary, 1)
                self.vacancy_stats[current_year] = CustomTuple(0, 0)
            else:
                self.years_stats[current_year].totalSalary += vacancy.salary
                self.years_stats[current_year].count += 1

            if vacancy.area_name not in self.cities_stats.keys():
                self.cities_stats[vacancy.area_name] = CustomTuple(vacancy.salary, 1)
            else:
                self.cities_stats[vacancy.area_name].totalSalary += vacancy.salary
                self.cities_stats[vacancy.area_name].count += 1

            if self.profession in vacancy.name:
                self.vacancy_stats[current_year].totalSalary += vacancy.salary
                self.vacancy_stats[current_year].count += 1

    def equalize_statistic(self) -> None:
        """Метод нормировки статистики в конкретном словаре

        :return: nothing
        """
        for year in self.years_stats.keys():
            self.years_stats[year].totalSalary = \
                int(self.years_stats[year].totalSalary //
                    self.years_stats[year].count)

        list_for_deleting = list()
        for city in self.cities_stats.keys():
            percent_count = round(self.cities_stats[city].count / self.cities_count, 4)
            if percent_count < 0.01:
                list_for_deleting.append(city)
            else:
                self.cities_stats[city].totalSalary = \
                    int(self.cities_stats[city].totalSalary //
                        self.cities_stats[city].count)
                self.cities_stats[city].count = percent_count

        for city in list_for_deleting:
            del [self.cities_stats[city]]
        for year in self.vacancy_stats.keys():
            if self.vacancy_stats[year].count != 0:
                self.vacancy_stats[year].totalSalary = \
                    int(self.vacancy_stats[year].totalSalary //
                        self.vacancy_stats[year].count)

    @staticmethod
    def print_first_string(string_for_output: str,
                           current_dictionary: dict,
                           value: str) -> None:
        """Метод для печати с правильным оформлением вывода всей статистики в консоль

        :param string_for_output: Строка на выход
        :param current_dictionary: Текущий обработанный словарь
        :param value: Конкретное значение для выборки

        :return: nothing
        """
        flag, index = False, 0
        print(string_for_output, end='')
        for year in current_dictionary.keys():
            if index == 0:
                print(' {', end='')
                index += 1
            printEnd = ', '
            if year == max(current_dictionary.keys()):
                printEnd = ''
                flag = True
            print(f'{year}: {getattr(current_dictionary[year], value)}', end=printEnd)
        if flag:
            print('}')

    @staticmethod
    def print_cities(string_for_output: str,
                     current_dictionary: dict,
                     names_list: list,
                     value: str) -> None:
        """Печать статистики зависящей от конкретного города

        :param string_for_output: Строка для вывода
        :param current_dictionary: Текущий словарь со статистикой
        :param names_list: Лист наименований
        :param value: Конкретное значение для выборки

        :return: nothing
        """
        flag, index = False, 0
        print(string_for_output, end='')
        for current_name in names_list:
            if index == 0:
                print(' {', end='')
            printEnd = ', '
            if index == len(names_list) - 1:
                printEnd = ''
                flag = True
            print(f"'{current_name}': {getattr(current_dictionary[current_name], value)}",
                  end=printEnd)
            index += 1
        if flag:
            print('}')

    def make_table(self):
        """Метод вызова всего необходимого для печати

        :return: nothing
        """
        self.calc(self.years_stats, "totalSalary")
        self.calc(self.years_stats, "count")
        self.calc(self.vacancy_stats, "totalSalary")
        self.calc(self.vacancy_stats, "count")
        # if len(data_vacancies) == 0:
        #    return {x: 0 for x in self.__list_years}
        cities_sorted = sorted(self.cities_stats, key=lambda x: self.cities_stats[x].totalSalary, reverse=True)
        del cities_sorted[10:]
        self.calc(self.cities_stats, "totalSalary")
        cities_sorted = sorted(self.cities_stats, key=lambda x: self.cities_stats[x].count, reverse=True)
        del cities_sorted[10:]
        self.calc(self.cities_stats, "count")
        self.list_of_all_dictionaries.insert(0, inserted_data.profession)
        print(self.list_of_all_dictionaries, end='\n', sep='\n\n')
        if self.word_for_choice.lower() == 'вакансии':
            self.report.generate_excel(*self.list_of_all_dictionaries)
        elif self.word_for_choice.lower() == 'статистика':
            self.report.generate_image(*self.list_of_all_dictionaries)
        else:
            print('Данные введены неправильно')

    def calc(self, dictionary: dict, value: str):
        """Метод вызволения словарей из объектов и добавления их в общий список

        :param dictionary: Конкретный объект;
        :param value: Конкретное значение для вызволения

        :return: nothing
        """
        common_vocabulary = dict()
        for year in dictionary.keys():
            common_vocabulary[year] = getattr(dictionary[year], value)
        self.list_of_all_dictionaries.append(common_vocabulary)


def __auto_width(ws):
    """
    Внешняя функции выравнивания по ширине Excel-таблички

    :param ws: Конкретная ячейка

    :return: nothing
    """
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length * 1.23


class Report:
    """Библиотека генерации файлов отчёта в виде .pdf .png .xlsx
    """

    @staticmethod
    def generate_pdf(input_name: str,
                     dynamics_slr: dict,
                     dynamics_count_vac: dict,
                     dynamics_slr_name: dict,
                     dynamics_count_vac_name: dict,
                     dynamics_slr_cities: dict,
                     dynamics_count_vac_cities: dict):
        """Метод генерации отчета в виде .pdf совмещающего и графики, и таблицы

        :param input_name: Название файла
        :param dynamics_slr: Словарь с годами и зарплатами
        :param dynamics_count_vac: Словарь с годами и количеством
        :param dynamics_slr_name: Словарь с наименованием и зарплатой
        :param dynamics_count_vac_name: Словарь с наименованием и количеством вакансий
        :param dynamics_slr_cities: Словарь заплаты по городам
        :param dynamics_count_vac_cities: Словарь с количеством вакансий по городам


        :return: nothing
        """
        headers1, headers2, headers3 = (["Год", "Средняя зарплата", f"Средняя зарплата - {input_name}",
                                         "Количество вакансий", f"Количество вакансий - {input_name}"],
                                        ["Город", "Уровень зарплат"], ["Город", "Доля вакансий"])
        rows1 = list(map(lambda year: [year] + [dictionary[year] for dictionary in
                                                (dynamics_slr, dynamics_count_vac,
                                                 dynamics_slr_name, dynamics_count_vac_name)], dynamics_slr.keys()))
        rows2 = list(map(lambda city: [city, dynamics_slr_cities[city]], dynamics_slr_cities.keys()))
        rows3 = list(map(lambda city: [city, dynamics_count_vac_cities[city]], dynamics_count_vac_cities.keys()))

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(graph_name='graph.png',
                                       vacancy_name=input_name, headers1=headers1, headers2=headers2,
                                       headers3=headers3,
                                       rows1=rows1, rows2=rows2, rows3=rows3)
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)

    @staticmethod
    def generate_image(input_name: str,
                       dynamics_slr: dict,
                       dynamics_count_vac: dict,
                       dynamics_slr_name: dict,
                       dynamics_count_vac_name: dict,
                       dynamics_slr_cities: dict,
                       dynamics_count_vac_cities: dict):
        """Метод генерации графиков отчета в .png

        :param input_name: Название файла
        :param dynamics_slr: Словарь с годами и зарплатами
        :param dynamics_count_vac: Словарь с годами и количеством
        :param dynamics_slr_name: Словарь с наименованием и зарплатой
        :param dynamics_count_vac_name: Словарь с наименованием и количеством вакансий
        :param dynamics_slr_cities: Словарь заплаты по городам
        :param dynamics_count_vac_cities: Словарь с количеством вакансий по городам


        :return: nothing
        """
        fig = plt.figure(figsize=(10, 6))
        plt.rcParams['font.size'] = '8'
        width = 0.4
        years = np.arange(len(dynamics_slr.keys()))
        ax = fig.add_subplot(221)
        ax.bar(years - width / 2,
               dynamics_slr.values(),
               width,
               label='средняя з/п')
        ax.bar(years + width / 2,
               dynamics_slr_name.values(),
               width,
               label=f'з/п {input_name}')
        ax.set_title('Уровень зарплат по годам')
        ax.set_xticks(years, dynamics_slr.keys(), rotation='vertical')
        ax.set_xticklabels(dynamics_slr.keys())
        ax.legend()

        bx = fig.add_subplot(222)
        bx.bar(years - width / 2,
               dynamics_count_vac.values(),
               width,
               label='Количество вакансий')
        bx.bar(years + width / 2,
               dynamics_count_vac_name.values(),
               width,
               label=f'Количество вакансий\n{input_name}')
        bx.set_title('Количество вакансий по годам')
        bx.set_xticks(years, dynamics_slr.keys(), rotation='vertical')
        bx.set_xticklabels(dynamics_slr.keys())
        bx.legend()
        bx.grid(axis='y')

        dynamics_slr_cities_rev = dict(reversed(list(dynamics_slr_cities.items())[:10]))
        cities_slr = np.arange(len(dynamics_slr_cities_rev.keys()))
        cx = fig.add_subplot(223)
        cx.barh(cities_slr - width / 2, dynamics_slr_cities_rev.values(), width + 0.2)
        cx.set_title('Уровень зарплат по годам')
        cx.set_yticks(cities_slr)
        cx.set_yticklabels(dynamics_slr_cities_rev.keys())
        cx.grid(axis='x')

        dx = fig.add_subplot(224)
        dynamics_count_vac_cit_rev = dict(reversed(list(dynamics_count_vac_cities.items())))
        dx.pie(dynamics_count_vac_cit_rev.values(),
               labels=dynamics_count_vac_cit_rev.keys())
        dx.set_title('Доля вакансий по городам')
        dx.axis('equal')
        fig.tight_layout()

        fig.savefig('graph.png')

    def generate_excel(self,
                       input_name: str,
                       dynamics_slr: dict,
                       dynamics_count_vac: dict,
                       dynamics_slr_name: dict,
                       dynamics_count_vac_name: dict,
                       dynamics_slr_cities: dict,
                       dynamics_count_vac_cities: dict):
        """Генерация XLSX файла отчёта

        :param input_name: Название файла
        :param dynamics_slr: Словарь с годами и зарплатами
        :param dynamics_count_vac: Словарь с годами и количеством
        :param dynamics_slr_name: Словарь с наименованием и зарплатой
        :param dynamics_count_vac_name: Словарь с наименованием и количеством вакансий
        :param dynamics_slr_cities: Словарь заплаты по городам
        :param dynamics_count_vac_cities: Словарь с количеством вакансий по городам

        :return: nothing
        """
        workbook = Workbook()
        stats_by_year = workbook.worksheets[0]
        stats_by_year.title = "Cтатистика по годам"
        stats_by_city = workbook.create_sheet("Cтатистика по городам")

        stats_by_year.append(['Год', 'Средняя зарплата', f"Средняя зарплата - {input_name}",
                              'Количество вакансий', f"Количество вакансий - {input_name}"])
        for i, year in enumerate(dynamics_slr.keys(), 2):
            stats_by_year.cell(row=i, column=1, value=year)
            for j, dictionary in enumerate((dynamics_slr, dynamics_count_vac,
                                            dynamics_slr_name, dynamics_count_vac_name), 2):
                stats_by_year.cell(row=i, column=j, value=dictionary[year])

        stats_by_city.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for i, city in enumerate(dynamics_slr_cities.keys(), 2):
            stats_by_city.cell(row=i, column=1, value=city)
            stats_by_city.cell(
                row=i, column=2, value=dynamics_slr_cities[city])
        for i, city in enumerate(dynamics_count_vac_cities.keys(), 2):
            stats_by_city.cell(row=i, column=4, value=city)
            stats_by_city.cell(row=i, column=5, value=dynamics_count_vac_cities[city])

        self.workbook(workbook)
        workbook.save('report.xlsx')

    @staticmethod
    def workbook(wb):
        """Нормировка конкретизированного workbook под свои задачи

        :param wb: Конкретная ячейка

        :return: nothing
        """
        bold_font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        outline = Border(top=thin, left=thin, right=thin, bottom=thin)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "")
                             for cell in column_cells)
                worksheet.column_dimensions[column_cells[0]
                                            .column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value is None:
                    continue
                for cell in column:
                    cell.border = outline


inserted_data = InputConnect()
inserted_data.start_entering()
current_dataset = DataSet(inserted_data.file_name, list())
current_dataset.put_vacancies()
inserted_data.count_vacancies(current_dataset.vacancies_objects)
inserted_data.equalize_statistic()
inserted_data.make_table()
